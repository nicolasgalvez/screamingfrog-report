"""Parse Screaming Frog CSV exports and generate an Excel report."""

import re
from collections import defaultdict
from pathlib import Path
from urllib.parse import urlparse

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# SF uses "Address" for the URL column in bulk exports
ACCESSIBILITY_URL_COLUMNS = ("Address", "URL")

# Styling
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
LINK_FONT = Font(color="0563C1", underline="single")
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")

# Excel sheet name limits
MAX_SHEET_NAME = 31
INVALID_SHEET_CHARS = re.compile(r"[\\/*?\[\]:]")

# Unified per-page columns
PAGE_COLUMNS = [
    "Type",
    "Issue",
    "Priority",
    "Details",
    "Description",
    "How To Fix",
    "Help URL",
]


# ---------------------------------------------------------------------------
# URL helpers
# ---------------------------------------------------------------------------

_URL_RE = re.compile(r"^https?://", re.IGNORECASE)


def _normalize_url(url: str) -> str:
    return url.replace("http://", "https://", 1) if url.startswith("http://") else url


def _is_url(value: str) -> bool:
    return bool(_URL_RE.match(value))


def _clean(value) -> str:
    """Convert a value to string, replacing NaN/None with empty string."""
    if pd.isna(value):
        return ""
    s = str(value)
    return "" if s.lower() == "nan" else s


# ---------------------------------------------------------------------------
# CSV loaders
# ---------------------------------------------------------------------------


def _find_csv(export_dir: Path, pattern: str) -> Path | None:
    matches = sorted(export_dir.glob(pattern))
    return matches[0] if matches else None


def _load_issues_overview(export_dir: Path) -> pd.DataFrame | None:
    csv_path = _find_csv(export_dir, "*ssues*verview*.csv")
    if csv_path is None:
        csv_path = _find_csv(export_dir, "*issues*.csv")
    if csv_path is None:
        return None
    df = pd.read_csv(csv_path, encoding="utf-8-sig")
    if "Issue Name" not in df.columns:
        return None
    return df


def _load_accessibility(export_dir: Path) -> pd.DataFrame | None:
    csv_path = _find_csv(export_dir, "*all_violations*.csv")
    if csv_path is None:
        csv_path = _find_csv(export_dir, "*ccessibility*iolation*.csv")
    if csv_path is None:
        return None
    df = pd.read_csv(csv_path, encoding="utf-8-sig")
    url_col = next((c for c in df.columns if c in ACCESSIBILITY_URL_COLUMNS), None)
    if url_col is None:
        return None
    if url_col != "URL":
        df = df.rename(columns={url_col: "URL"})
    df["URL"] = df["URL"].apply(_normalize_url)
    return df


def _load_accessibility_summary(export_dir: Path) -> pd.DataFrame | None:
    csv_path = _find_csv(export_dir, "*ccessibility*ummary*.csv")
    if csv_path is None:
        return None
    return pd.read_csv(csv_path, encoding="utf-8-sig")


_ASSET_EXT_RE = re.compile(
    r"\.(?:jpg|jpeg|png|gif|svg|webp|ico|bmp|tiff"
    r"|pdf|doc|docx|xls|xlsx|ppt|pptx"
    r"|css|js|json|xml|txt|csv"
    r"|woff|woff2|ttf|eot|otf"
    r"|mp4|mp3|wav|avi|mov|webm"
    r"|zip|gz|tar|rar)(?:\?.*)?$",
    re.IGNORECASE,
)


def _load_internal_urls(export_dir: Path) -> set[str]:
    """Load the set of internal HTML page URLs from the Internal:All export."""
    csv_path = _find_csv(export_dir, "*internal_all*.csv")
    if csv_path is None:
        csv_path = _find_csv(export_dir, "*nternal*ll*.csv")
    if csv_path is None:
        return set()
    df = pd.read_csv(
        csv_path, encoding="utf-8-sig", usecols=["Address", "Content Type"]
    )
    # Only include HTML pages (not images, PDFs, CSS, JS, etc.)
    html_mask = df["Content Type"].str.contains("html", na=False)
    # Exclude asset file extensions even if SF reports them as text/html (soft 404s)
    not_asset = ~df["Address"].str.contains(_ASSET_EXT_RE, na=False)
    return {
        _normalize_url(u) for u in df.loc[html_mask & not_asset, "Address"].dropna()
    }


def _load_per_page_issues(
    export_dir: Path,
    issues_overview: pd.DataFrame | None,
    internal_urls: set[str],
) -> dict[str, list[dict]]:
    """Load per-page issues from issues_reports/ folder.

    Only includes URLs in the internal_urls set.
    Returns {url: [row_dict, ...]} where each row_dict has PAGE_COLUMNS keys.
    """
    issues_dir = export_dir / "issues_reports"
    if not issues_dir.is_dir():
        return {}

    # Build lookup from Issues Overview for metadata
    overview_lookup: dict[str, dict] = {}
    if issues_overview is not None:
        for _, row in issues_overview.iterrows():
            name = str(row.get("Issue Name", ""))
            overview_lookup[name.lower()] = {
                "Priority": _clean(row.get("Issue Priority", "")),
                "Description": _clean(row.get("Description", "")),
                "How To Fix": _clean(row.get("How To Fix", "")),
                "Help URL": _clean(row.get("Help URL", "")),
                "Issue Type": _clean(row.get("Issue Type", "")),
            }

    url_issues: dict[str, list[dict]] = defaultdict(list)

    for csv_path in sorted(issues_dir.glob("*.csv")):
        # Derive issue name from filename
        issue_name = csv_path.stem.replace("_", " ").title()

        # Skip inlinks CSVs — they're about links pointing TO the issue, not the page
        if "inlinks" in csv_path.stem.lower():
            continue

        try:
            df = pd.read_csv(csv_path, encoding="utf-8-sig")
        except Exception:
            continue

        # Find the URL/Address column
        addr_col = next(
            (c for c in df.columns if c in ("Address", "URL")),
            None,
        )
        if addr_col is None:
            continue

        # Try to match to Issues Overview for richer metadata
        meta = _match_overview(issue_name, overview_lookup)

        for _, row in df.iterrows():
            url = _normalize_url(str(row[addr_col]))
            # Only include internal URLs
            if internal_urls and url not in internal_urls:
                continue

            # Collect any extra detail columns (skip Address and common noise)
            skip = {addr_col, "Indexability", "Indexability Status"}
            details = []
            for col in df.columns:
                if col in skip:
                    continue
                val = row[col]
                if pd.notna(val) and str(val).strip():
                    details.append(f"{col}: {val}")
            detail_str = "; ".join(details) if details else ""

            url_issues[url].append(
                {
                    "Type": "Issue",
                    "Issue": meta.get("name", issue_name),
                    "Priority": meta.get("Priority", ""),
                    "Details": detail_str,
                    "Description": meta.get("Description", ""),
                    "How To Fix": meta.get("How To Fix", ""),
                    "Help URL": meta.get("Help URL", ""),
                }
            )

    return dict(url_issues)


def _match_overview(filename_issue: str, overview: dict[str, dict]) -> dict[str, str]:
    """Try to fuzzy-match an issue filename to an Issues Overview entry."""
    # Exact lowercase match
    key = filename_issue.lower()
    if key in overview:
        return {**overview[key], "name": filename_issue}

    # Try matching by checking if overview name words are in filename
    fn_words = set(key.split())
    best_score = 0
    best_match: dict[str, str] = {}
    for ov_key, ov_meta in overview.items():
        ov_words = set(ov_key.split())
        # Skip accessibility issues — those come from the violations CSV
        if ov_key.startswith("accessibility:"):
            continue
        overlap = len(fn_words & ov_words)
        if overlap > best_score and overlap >= 2:
            best_score = overlap
            best_match = {**ov_meta, "name": ov_key.title()}

    return best_match if best_match else {"name": filename_issue}


# ---------------------------------------------------------------------------
# Accessibility → unified rows
# ---------------------------------------------------------------------------


def _a11y_to_rows(a11y_df: pd.DataFrame, url: str) -> list[dict]:
    """Convert accessibility violations for a URL into unified row dicts."""
    page_df = a11y_df[a11y_df["URL"] == url]
    rows = []
    for _, row in page_df.iterrows():
        rows.append(
            {
                "Type": "Accessibility",
                "Issue": _clean(row.get("Issue", "")),
                "Priority": _clean(row.get("Priority", "")),
                "Details": _clean(row.get("Location on Page", "")),
                "Description": _clean(row.get("Issue Description", "")),
                "How To Fix": _clean(row.get("How To Fix", "")),
                "Help URL": _clean(row.get("Help URL", "")),
            }
        )
    return rows


# ---------------------------------------------------------------------------
# Sheet helpers
# ---------------------------------------------------------------------------


def _url_to_sheet_name(url: str, seen: dict[str, int]) -> str:
    parsed = urlparse(url)
    path = parsed.path.strip("/")
    name = path.replace("/", " - ") if path else "home"
    name = INVALID_SHEET_CHARS.sub("", name).strip() or "page"

    if len(name) > MAX_SHEET_NAME - 4:
        name = name[: MAX_SHEET_NAME - 4].rstrip(" -")

    base = name
    if base in seen:
        seen[base] += 1
        suffix = f" ({seen[base]})"
        name = name[: MAX_SHEET_NAME - len(suffix)] + suffix
    else:
        seen[base] = 0

    return name


def _style_header_row(ws, row_num: int) -> None:
    for cell in ws[row_num]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_ALIGNMENT


def _auto_column_widths(ws, max_width: int = 60) -> None:
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = 0
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        width = min(max_len + 2, max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(width, 12)


def _make_urls_clickable(ws) -> None:
    """Scan all cells and make any URL values into clickable hyperlinks."""
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and _is_url(cell.value):
                cell.hyperlink = cell.value
                cell.font = LINK_FONT


def _df_to_sheet(wb: Workbook, name: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(title=name)
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append([str(v) if pd.notna(v) else "" for v in row])
    _style_header_row(ws, 1)
    _make_urls_clickable(ws)
    _auto_column_widths(ws)
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Main report generation
# ---------------------------------------------------------------------------


def generate_report(export_dir: Path, output_path: Path) -> Path:
    """Generate an Excel report from Screaming Frog CSV exports.

    Produces a workbook with:
    - Issues Summary sheet
    - Accessibility Summary sheet
    - Pages index sheet (with hyperlinks)
    - One sheet per page with combined issues + accessibility violations
    """
    wb = Workbook()
    wb.remove(wb.active)

    # --- Issues Summary ---
    issues_df = _load_issues_overview(export_dir)
    if issues_df is not None:
        _df_to_sheet(wb, "Issues Summary", issues_df)
        print(f"  Issues Summary: {len(issues_df)} issue types")
    else:
        print("  Warning: No Issues Overview CSV found, skipping sheet.")

    # --- Accessibility Summary ---
    a11y_summary_df = _load_accessibility_summary(export_dir)
    a11y_df = _load_accessibility(export_dir)

    if a11y_summary_df is not None:
        _df_to_sheet(wb, "Accessibility Summary", a11y_summary_df)
        print(f"  Accessibility Summary: {len(a11y_summary_df)} rows (from SF report)")
    elif a11y_df is not None:
        issue_col = next(
            (
                c
                for c in a11y_df.columns
                if c.lower() in ("issue", "violation", "issue name")
            ),
            a11y_df.columns[1] if len(a11y_df.columns) > 1 else None,
        )
        if issue_col:
            summary = (
                a11y_df.groupby(issue_col)
                .agg(Count=("URL", "size"), Pages_Affected=("URL", "nunique"))
                .sort_values("Count", ascending=False)
                .reset_index()
            )
            _df_to_sheet(wb, "Accessibility Summary", summary)
            print(f"  Accessibility Summary: {len(summary)} violation types (computed)")

    # --- Load per-page issues (internal URLs only) ---
    internal_urls = _load_internal_urls(export_dir)
    per_page_issues = _load_per_page_issues(export_dir, issues_df, internal_urls)
    issue_url_count = len(per_page_issues)
    print(f"  Per-page issues loaded for {issue_url_count} URLs")

    # --- Filter accessibility to internal HTML pages only ---
    if a11y_df is not None and internal_urls:
        a11y_df = a11y_df[a11y_df["URL"].isin(internal_urls)]

    # --- Build combined per-page data ---
    # Collect all URLs that have either accessibility or issue data
    all_urls: set[str] = set()
    if a11y_df is not None:
        all_urls.update(a11y_df["URL"].unique())
    all_urls.update(per_page_issues.keys())

    # Build {url: [row_dicts]} combining both sources
    url_rows: dict[str, list[dict]] = {}
    for url in all_urls:
        rows: list[dict] = []
        if a11y_df is not None:
            rows.extend(_a11y_to_rows(a11y_df, url))
        rows.extend(per_page_issues.get(url, []))
        if rows:
            url_rows[url] = rows

    # --- Deduplicate pages with identical combined data ---
    def _fingerprint(rows: list[dict]) -> str:
        sortable = [tuple(sorted(r.items())) for r in rows]
        sortable.sort()
        return repr(sortable)

    url_fingerprints = {url: _fingerprint(rows) for url, rows in url_rows.items()}
    fp_to_urls: dict[str, list[str]] = defaultdict(list)
    for url, fp in sorted(url_fingerprints.items()):
        fp_to_urls[fp].append(url)

    # --- Write per-page sheets ---
    page_index: list[dict] = []
    sheet_names_seen: dict[str, int] = {}

    for fp, urls in fp_to_urls.items():
        representative = urls[0]
        rows = url_rows[representative]
        sheet_name = _url_to_sheet_name(representative, sheet_names_seen)

        ws = wb.create_sheet(title=sheet_name)

        # URLs at top
        if len(urls) == 1:
            ws.append(["URL", urls[0]])
            cell = ws.cell(row=1, column=2)
            cell.hyperlink = urls[0]
            cell.font = LINK_FONT
        else:
            ws.append(["URLs", f"{len(urls)} pages with identical issues"])
            for u in urls:
                ws.append(["", u])
                cell = ws.cell(row=ws.max_row, column=2)
                cell.hyperlink = u
                cell.font = LINK_FONT

        ws.append([])  # separator
        ws["A1"].font = Font(bold=True)

        # Data table
        header_row = ws.max_row + 1
        ws.append(PAGE_COLUMNS)
        for r in rows:
            ws.append([str(r.get(c, "")) for c in PAGE_COLUMNS])

        # Make Help URL column clickable
        help_col_idx = PAGE_COLUMNS.index("Help URL") + 1
        for row_num in range(header_row + 1, ws.max_row + 1):
            cell = ws.cell(row=row_num, column=help_col_idx)
            if cell.value and _is_url(str(cell.value)):
                cell.hyperlink = cell.value
                cell.font = LINK_FONT

        _style_header_row(ws, header_row)
        _auto_column_widths(ws)
        ws.freeze_panes = f"A{header_row + 1}"

        # Count by type for the index
        a11y_count = sum(1 for r in rows if r["Type"] == "Accessibility")
        issue_count = sum(1 for r in rows if r["Type"] == "Issue")

        for u in urls:
            page_index.append(
                {
                    "URL": u,
                    "Sheet": sheet_name,
                    "Accessibility": a11y_count,
                    "Issues": issue_count,
                    "Duplicates": len(urls) if len(urls) > 1 else "",
                }
            )

    # --- Pages index with hyperlinks ---
    ws = wb.create_sheet(title="Pages")
    headers = ["URL", "Sheet", "Accessibility", "Issues", "Duplicates"]
    ws.append(headers)
    for entry in page_index:
        row_num = ws.max_row + 1
        dupes = entry["Duplicates"]
        ws.append(
            [
                entry["URL"],
                entry["Sheet"],
                str(entry["Accessibility"]),
                str(entry["Issues"]),
                str(dupes) if dupes else "",
            ]
        )
        # Clickable URL
        url_cell = ws.cell(row=row_num, column=1)
        url_cell.hyperlink = entry["URL"]
        url_cell.font = LINK_FONT
        # Clickable sheet link
        sheet_ref = entry["Sheet"].replace("'", "''")
        sheet_cell = ws.cell(row=row_num, column=2)
        sheet_cell.hyperlink = f"#'{sheet_ref}'!A1"
        sheet_cell.font = LINK_FONT

    _style_header_row(ws, 1)
    _auto_column_widths(ws)
    ws.freeze_panes = "A2"

    # Move Pages to after summaries
    summary_count = sum(
        1
        for name in wb.sheetnames
        if name in ("Issues Summary", "Accessibility Summary")
    )
    wb.move_sheet("Pages", offset=-(len(wb.sheetnames) - 1 - summary_count))

    unique_sheets = len(fp_to_urls)
    total_pages = len(page_index)
    dupes = total_pages - unique_sheets
    print(f"  Pages index: {total_pages} pages")
    print(f"  Per-page sheets: {unique_sheets} unique ({dupes} duplicates collapsed)")

    if len(wb.sheetnames) == 0:
        raise RuntimeError(
            f"No valid CSVs found in {export_dir}. "
            "Expected Issues Overview and/or Accessibility Violations exports."
        )

    wb.save(str(output_path))
    print(f"\nReport saved to {output_path}")
    return output_path
